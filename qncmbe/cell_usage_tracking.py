'''
This is a tool for estimating effusion cell usage in the QNC-MBE system

Almost everything is contained in the CellUsageCalculator class.
The basic idea is to pull cell temperature data, as stored by Molly, and to pull A,B,C calibration coefficients from the Excel document which we manually update.

From this, we can estimate the cell usage as a function of time. There are a number of uncertainties though. 
This relies on accurate tracking of the A,B,C coefficients which may not always be true.

More importantly, when we calculate A,B,C coefficients, we calculate them to find the flux hitting the centre of the wafer.
To calculate the element usage, though, we need to know the total number of atoms leaving the cell.

For now, we assume that for SUMO cells (Ga, In), 3.5% of the total flux hits the wafer, while for conical cells (Al) 7% of the total flux hits the wafer.
This was determined approximately in 2017 by comparing the estimated usage to the actual usage.
However, it would be worth doing more accurate measurements in the future.

See example file for example of usage.
'''

# Standard library imports (not included in setup.py)
import os
import datetime as dt

# Non-standard library imports (included in setup.py)
import openpyxl as xl
import numpy as np
from scipy.integrate import cumtrapz
import matplotlib.dates as mdates

# qncmbe imports
import qncmbe.data_import.data_import_utils as datimp

try:
    # To deal with the error message
    # "FutureWarning: Using an implicitly registered datetime converter for a matplotlib plotting method. The converter was registered by pandas on import. Future versions of pandas will require you to explicitly register matplotlib converters."
    # Following is the recommended solution.
    # Presumably this is not an issue if pandas is not installed, so I put this in a try block
    from pandas.plotting import register_matplotlib_converters
    register_matplotlib_converters()
except:
    pass

valid_cells = ['Ga1','Ga2','In1','In2','Al1']

avogadro = 6.022140857e23
wafer_radius = 3.81
wafer_area = np.pi*wafer_radius**2

atomic_mass = {
    'Al1': 26.9815,
    'Ga1': 69.723,
    'Ga2': 69.723,
    'In1': 114.818,
    'In2': 114.818
}

# Estimated fraction of the beam that hits a 3" wafer
beam_efficiency = {
    'Al1': 0.070,
    'Ga1': 0.035,
    'Ga2': 0.035,
    'In1': 0.035,
    'In2': 0.035
}

class CellUsageCalculator():
    def __init__(self, start_date, end_date, cells, cell_pars_file, save_dir = '.\\saved_cell_data', delta_t = 300, regen_data = False):
        '''
        - start_date and end_date should be strings of the form 'YYYY-MM-DD'
        - cells should be a list of cells (e.g., ['Ga1','Ga2','Al1'])
        - cell_pars_file should be the full filepath to the Excel file containing the ABC parameters for each cell
        - save_dir is the directory in which partial data will be saved
            (To avoid collecting enormous amounts of data from Molly at once (takes a very long time), data is loaded one day at a time and saved into save_dir.)
        - delta_t is the spacing (s) between data samples. Default is 300 s (5 min)
        - regen_data determines whether or not to regenerate the Cell_data_yyyy-mm-dd.csv files. Should set this to True if, e.g., you've added an additional cell since the last run
        '''

        fmt_str = '%Y-%m-%d'

        self.start_date = dt.datetime.strptime(start_date, fmt_str)
        self.end_date = dt.datetime.strptime(end_date, fmt_str)

        if self.start_date >= self.end_date:
            raise Exception("start_date must be before end_date!")

        self.cell_pars_file = cell_pars_file
        self.save_dir = save_dir

        self.delta_t = delta_t

        self.regen_data = regen_data


        self.cells = cells
        if not set(self.cells).issubset(set(valid_cells)):
            raise Exception(f"Invalid cell selection. Only allowed values are {valid_cells}")


        self.value_names_list = ['Molly time']

        for cell in self.cells:
            self.value_names_list.append(f'{cell} base measured')

        self.time = np.zeros(0)
        self.temperature = {}

        self.A = {}
        self.B = {}
        self.C = {}

        self.mass_usage = {}
        self.particle_usage = {}

        
    def generate_temperature_data_files(self):
        '''
        Collect temperature data into .dat files.
        If self.regen_data == True, then it will not reuse existing datafiles and will regenerate them from Molly (might take a very long time!)
        '''
        day = self.start_date
        delta = dt.timedelta(days = 1)

        if not os.path.exists(self.save_dir):
            os.makedirs(self.save_dir)

        while day <= self.end_date:
            fname = f'cell_temperatures_{day.date()}.csv'
            fpath = os.path.join(self.save_dir, fname)

            write_data = True

            if os.path.exists(fpath):
                if self.regen_data:
                    print(f"Found temperature data for {day.date()} but overwriting it!")
                else: 
                    write_data = False
            else:
                print(f"No temperature data for {day.date()}. Generating it...")

            if write_data:
                
                data = datimp.get_data(
                            start_time = day,
                            end_time = day + delta,
                            value_names_list = self.value_names_list,
                            delta_t = self.delta_t,
                            interp = True)

                t = data['Molly time']

                T = {}


                for cell in self.cells:
                    T[cell] = data[f'{cell} base measured'] + 273.15



                out_arr = [t]
                header = 'Time (s)'
                for cell in self.cells:
                    out_arr.append(T[cell])
                    
                    header += f',{cell} temperature (K)'

                np.savetxt(fpath, np.stack(out_arr).transpose(), header = header, delimiter = ',')
            
            day += delta

    def collect_temperature_data(self):

        self.generate_temperature_data_files()

        day = self.start_date
        delta = dt.timedelta(days = 1)

        time_list = []
        temp_list = {cell: [] for cell in self.cells}

        while day <= self.end_date:
            fname = f'cell_temperatures_{day.date()}.csv'
            fpath = os.path.join(self.save_dir, fname)
            
            data = np.genfromtxt(fpath,names=True,delimiter=',')

            offset = (day - self.start_date).total_seconds()
            
            time_list.append(data['Time_s'] + offset)

            for cell in self.cells:
                try:
                    temp_list[cell].append(data[f'{cell}_temperature_K'])
                except ValueError:
                    raise(ValueError(f'Did not find cell {cell} in datafile. Might need to generate new datafiles by rerunning with regen_data = True.'))
            #for c in self.cells:
            #    print(data[f'{c}_temperature_K'].shape)

            day += delta
            
        self.time = np.concatenate(time_list)
        self.temperature = {cell: np.concatenate(temp_list[cell]) for cell in self.cells}

        # Discard unreasonable temperature values (set to zero Kelvin)
        # Sometimes if the signal is disconnected, an unrealistically large reading will be given,
        # which will completely invalidate any cell usage estimate
        T_max = 2500
        T_min = 0.0

        for cell in self.cells:
            mask = (self.temperature[cell] > T_max) | (self.temperature[cell] < T_min)

            self.temperature[cell][mask] = 0.0

    def get_time(self):

        if len(self.time) == 0:
            self.collect_temperature_data()

        return self.time

    def get_temperature(self):

        if not self.temperature:
            self.collect_temperature_data()

        return self.temperature

    def collect_ABC_coefs(self):

        t = self.get_time()

        for cell in self.cells:

            wb = xl.load_workbook(self.cell_pars_file, data_only = True)
            sheet = wb.get_sheet_by_name(cell)

            row = 10 # Starting row in the table

            if sheet.cell(row=row-2,column=2).value != "Growth #":
                print("WARNING: unexpected cell header value. Make sure the structure of the Calibration Parameters Excel sheet has not been modified." )

            A_raw = []
            B_raw = []
            C_raw = []
            t_raw = []

            break_count = 0

            while break_count < 20:
                t_val = sheet.cell(row = row, column = 3).value

                A_val = sheet.cell(row = row, column = 11).value
                B_val = sheet.cell(row = row, column = 12).value
                C_val = sheet.cell(row = row, column = 13).value

                if all([t_val, A_val, B_val, C_val]):
                    t_raw.append((t_val - self.start_date).total_seconds())
                    A_raw.append(A_val)
                    B_raw.append(B_val)
                    C_raw.append(C_val)

                    break_count = 0

                else:
                    break_count += 1

                row += 1

            t_raw = np.array(t_raw)

            sort_inds = np.argsort(t_raw)

            A_raw = np.array(A_raw)[sort_inds]
            B_raw = np.array(B_raw)[sort_inds]
            C_raw = np.array(C_raw)[sort_inds]

            self.A[cell] = np.interp(t, t_raw, A_raw)
            self.B[cell] = np.interp(t, t_raw, B_raw)
            self.C[cell] = np.interp(t, t_raw, C_raw)

    def get_ABC_coefs(self):
        
        if not self.A:
            self.collect_ABC_coefs()

        return self.A, self.B, self.C

    def calculate_element_usage(self):
        '''
        From the cell temperature data and the cell A,B,C coefficients,
        calculates the estimated usage of the cell in grams.

        Returns self.time (numpy array) and self.mass_usage (dictionary: one numpy array for each cell)
        '''
        
        if len(self.time) == 0:
            self.collect_temperature_data()

        if not self.A:
            self.collect_ABC_coefs()


        for cell in self.cells:

            A = self.A[cell]
            B = self.B[cell]
            C = self.C[cell]
            T = self.temperature[cell]

            t = self.time

            flux = 1e16*A*C*np.exp(-B/T)/np.sqrt(T)
            
            integrated_flux = cumtrapz(flux, t, initial = 0.0)
            
            self.particle_usage[cell] = wafer_area*integrated_flux/beam_efficiency[cell]
            
            self.mass_usage[cell] = self.particle_usage[cell]*atomic_mass[cell]/avogadro

    def get_mass_usage(self):
        '''
        Return usage as mass in (g). Returns a dictionary with one numpy array for each cell.
        '''

        if not self.mass_usage:
            self.calculate_element_usage()
        
        return self.mass_usage

    def get_particle_usage(self):
        '''
        Return usage as number of particles. Returns a dictionary with one numpy array for each cell.
        '''

        if not self.particle_usage:
            self.calculate_element_usage()

        return self.particle_usage

    def generate_usage_csv(self, fname = ''):

        t = self.get_time()
        m = self.get_mass_usage()
        N = self.get_particle_usage()

        if fname == '':
            fname = os.path.join(self.save_dir, 'cell_estimated_usage.csv')
        
        out_arr = [t]
        
        header = 'Time (s)'

        for cell in self.cells:
            out_arr.append(N[cell])
            out_arr.append(m[cell])
        
            header += f',{cell} usage (# of atoms),{cell} usage (g)'


        print(f"Saving usage data to {fname}")
        np.savetxt(fname, np.stack(out_arr).transpose(), header = header, delimiter = ',', comments='')


    def plot_temperatures(self, fig, ax, use_date_format = True):

        T = self.get_temperature()
        T_celsius = {cell: T[cell] - 273.15 for cell in self.cells}

        plot_cell_val(fig, ax, 
                      t = self.get_time(),
                      val = T_celsius,
                      cells = self.cells, 
                      start_date = self.start_date, 
                      use_date_format=use_date_format)

        ax.set_ylabel('Temperature (°C)')
        ax.set_title('Cell temperatures')

    def plot_mass_usage(self, fig, ax, use_date_format = True):

        plot_cell_val(fig, ax, 
                      t = self.get_time(),
                      val = self.get_mass_usage(),
                      cells = self.cells, 
                      start_date = self.start_date, 
                      use_date_format=use_date_format)

        ax.set_ylabel('Mass (g)')
        ax.set_title('Cell usage')

    def plot_particle_usage(self, fig, ax, use_date_format = True):

        plot_cell_val(fig, ax, 
                      t = self.get_time(),
                      val = self.get_particle_usage(),
                      cells = self.cells, 
                      start_date = self.start_date, 
                      use_date_format=use_date_format)

        ax.set_ylabel('Number of atoms')
        ax.set_title('Cell usage')

    def print_total_usage(self):

        header = f'Total element usage ({self.start_date.date()} to {self.end_date.date()})'

        string = '='*len(header)
        string += f'\n{header}\n'
        string += '='*len(header)

        m = self.get_mass_usage()
        N = self.get_particle_usage()

        for cell in self.cells:
            string += f'\n{cell}: {m[cell][-1]:7.2f} g   ({N[cell][-1]:.3e} atoms)'
        
        string += '\n'
        string += '='*len(header)

        print(string)
        return string


def plot_cell_val(fig, ax, t, val, cells, start_date, use_date_format = True):
    '''
    General function for plotting cell values.
    fig, ax should be matplotlib Figure and Axes objects.
    t should be a numpy array of time values
    val should be a dictionary containing one numpy array of data for each cell
    use_date determines whether to plot time on the x-axis as a date, or as number of days from start_date. Must provide start date if plotting in date format.
    start_date should be a datetime object
    '''

    if use_date_format:
        t_plt = np.datetime64(start_date.strftime("%Y-%m-%dT%H:%M:%S")) + t*np.timedelta64(1,'s')
    else:
        t_plt = t/86400

    for cell in cells:
        ax.plot(t_plt, val[cell], label = cell)

    ax.legend()

    if use_date_format:
        ax.set_xlabel('')
        ax.format_xdata = mdates.DateFormatter('%Y-%m-%d %H:%M:%S')
        fig.autofmt_xdate()
    else:
        ax.set_xlabel(f'Time (days from {start_date.date()})')
